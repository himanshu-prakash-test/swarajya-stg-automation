"""
Excel Reader Utility for the Swarajya Automation Framework.

Provides functions to:
1. Read login credentials from credentials.xlsx
2. Read test cases from login_test_cases.xlsx
3. Update test case results after execution (Automation_Result, Timestamp, Remarks)

Uses openpyxl for Excel I/O.
"""
import os
import logging
from datetime import datetime
from typing import Dict, List, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill

logger = logging.getLogger(__name__)

# ── Paths ──
_BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEST_DATA_DIR = os.path.join(_BASE_DIR, "test_data")
CREDENTIALS_FILE = os.path.join(TEST_DATA_DIR, "credentials.xlsx")
TEST_CASES_FILE = os.path.join(TEST_DATA_DIR, "login_test_cases.xlsx")


# ═══════════════════════════════════════════════════════════════════
# CREDENTIALS READER
# ═══════════════════════════════════════════════════════════════════

def read_credentials(role: str) -> Dict[str, str]:
    """
    Read login credentials for a specific role from credentials.xlsx.

    Args:
        role: The role to fetch credentials for (e.g., 'Employee', 'Manager').

    Returns:
        Dict with keys: 'role', 'employee_id', 'password', 'auth_code', 'is_valid'

    Raises:
        FileNotFoundError: If credentials.xlsx is missing.
        ValueError: If the specified role is not found.
    """
    if not os.path.exists(CREDENTIALS_FILE):
        raise FileNotFoundError(
            f"Credentials file not found: {CREDENTIALS_FILE}\n"
            "Please create it with columns: Role, Employee_ID, Password, Auth_Code, Is_Valid"
        )

    wb = openpyxl.load_workbook(CREDENTIALS_FILE, read_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    logger.info("Reading credentials for role: %s", role)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        row_dict = dict(zip(headers, row))
        if row_dict.get("Role", "").strip().lower() == role.strip().lower():
            wb.close()
            return {
                "role": row_dict["Role"].strip(),
                "employee_id": str(row_dict["Employee_ID"]).strip(),
                "password": str(row_dict["Password"]).strip(),
                "auth_code": str(row_dict["Auth_Code"]).strip(),
                "is_valid": str(row_dict.get("Is_Valid", "")).strip().lower() == "yes",
            }

    wb.close()
    raise ValueError(f"Role '{role}' not found in credentials.xlsx")


def read_all_credentials() -> List[Dict[str, str]]:
    """
    Read all credentials from credentials.xlsx.

    Returns:
        List of credential dicts for all roles.
    """
    if not os.path.exists(CREDENTIALS_FILE):
        raise FileNotFoundError(f"Credentials file not found: {CREDENTIALS_FILE}")

    wb = openpyxl.load_workbook(CREDENTIALS_FILE, read_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    credentials = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if any(v is not None for v in row):
            row_dict = dict(zip(headers, row))
            credentials.append({
                "role": str(row_dict.get("Role", "")).strip(),
                "employee_id": str(row_dict.get("Employee_ID", "")).strip(),
                "password": str(row_dict.get("Password", "")).strip(),
                "auth_code": str(row_dict.get("Auth_Code", "")).strip(),
                "is_valid": str(row_dict.get("Is_Valid", "")).strip().lower() == "yes",
            })

    wb.close()
    logger.info("Loaded credentials for %d roles", len(credentials))
    return credentials


# ═══════════════════════════════════════════════════════════════════
# TEST CASE READER
# ═══════════════════════════════════════════════════════════════════

def read_test_cases(
    sheet_name: Optional[str] = None,
    role_filter: Optional[str] = None,
    tc_id_filter: Optional[str] = None,
) -> List[Dict[str, str]]:
    """
    Read test cases from login_test_cases.xlsx.

    Args:
        sheet_name: 'Positive_Flows', 'Negative_Flows', or None for both.
        role_filter: Filter by role (e.g., 'Employee', 'Manager', 'General').
        tc_id_filter: Filter by specific Test Case ID.

    Returns:
        List of test case dicts with keys matching Excel headers.
    """
    if not os.path.exists(TEST_CASES_FILE):
        raise FileNotFoundError(f"Test cases file not found: {TEST_CASES_FILE}")

    wb = openpyxl.load_workbook(TEST_CASES_FILE, read_only=True)
    sheets = [sheet_name] if sheet_name else wb.sheetnames
    test_cases = []

    for sname in sheets:
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        headers = [cell.value for cell in ws[1]]

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if not any(v is not None for v in row):
                continue
            row_dict = dict(zip(headers, row))

            # Apply filters
            if role_filter:
                row_role = str(row_dict.get("Role", "")).strip().lower()
                if row_role != role_filter.strip().lower() and row_role != "general":
                    continue

            if tc_id_filter:
                if str(row_dict.get("Test Case ID", "")).strip() != tc_id_filter:
                    continue

            row_dict["_sheet"] = sname
            test_cases.append(row_dict)

    wb.close()
    logger.info(
        "Loaded %d test cases (sheet=%s, role=%s)",
        len(test_cases), sheet_name, role_filter,
    )
    return test_cases


# ═══════════════════════════════════════════════════════════════════
# TEST RESULT WRITER
# ═══════════════════════════════════════════════════════════════════

# Styling for result cells
_PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_PASS_FONT = Font(color="006100", bold=True)
_FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_FAIL_FONT = Font(color="9C0006", bold=True)
_SKIP_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_SKIP_FONT = Font(color="9C6500", bold=True)


def update_test_result(
    tc_id: str,
    result: str,
    remarks: str = "",
):
    """
    Update the Automation_Result, Execution_Timestamp, and Remarks columns
    for a specific Test Case ID in login_test_cases.xlsx.

    Args:
        tc_id: The Test Case ID (e.g., 'TC_EMP_001').
        result: 'PASS', 'FAIL', 'SKIPPED', or 'BLOCKED'.
        remarks: Optional remarks/failure reason.
    """
    if not os.path.exists(TEST_CASES_FILE):
        logger.warning("Test cases file not found, skipping result update")
        return

    wb = openpyxl.load_workbook(TEST_CASES_FILE)
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]

        # Find column indices for result columns
        tc_col = headers.index("Test Case ID") + 1 if "Test Case ID" in headers else None
        result_col = headers.index("Automation_Result") + 1 if "Automation_Result" in headers else None
        ts_col = headers.index("Execution_Timestamp") + 1 if "Execution_Timestamp" in headers else None
        remarks_col = headers.index("Remarks") + 1 if "Remarks" in headers else None

        if not tc_col or not result_col:
            continue

        for row_idx in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=row_idx, column=tc_col).value
            if str(cell_val).strip() == tc_id:
                # Write result
                result_cell = ws.cell(row=row_idx, column=result_col, value=result)
                if result == "PASS":
                    result_cell.fill = _PASS_FILL
                    result_cell.font = _PASS_FONT
                elif result == "FAIL":
                    result_cell.fill = _FAIL_FILL
                    result_cell.font = _FAIL_FONT
                elif result in ("SKIPPED", "BLOCKED"):
                    result_cell.fill = _SKIP_FILL
                    result_cell.font = _SKIP_FONT

                # Write timestamp
                if ts_col:
                    ws.cell(row=row_idx, column=ts_col, value=timestamp)

                # Write remarks
                if remarks_col and remarks:
                    ws.cell(row=row_idx, column=remarks_col, value=remarks)

                logger.info("Updated TC %s: %s", tc_id, result)
                break

    try:
        wb.save(TEST_CASES_FILE)
        wb.close()
    except Exception as e:
        logger.warning("Could not save Excel test results to file (may be locked/open): %s", e)
