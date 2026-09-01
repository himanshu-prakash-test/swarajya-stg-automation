"""Workbook reader and result writer for employee-management tests."""

import os
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

LOCAL_WORKBOOK = os.path.join(
    os.path.dirname(__file__),
    "test_data",
    "Swarajya-Update-Employee-test-cases.xlsx",
)
DOWNLOADS_WORKBOOK = "/Users/mrugankkapse/Downloads/Swarajya-Update-Employee-test-cases.xlsx"
_PASS = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_FAIL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_SKIP = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")


def workbook_path():
    configured = os.environ.get("EMPLOYEE_UPDATE_WORKBOOK")
    if configured:
        return configured
    if os.path.exists(LOCAL_WORKBOOK):
        return LOCAL_WORKBOOK
    return DOWNLOADS_WORKBOOK


def read_employee_cases():
    path = workbook_path()
    if not os.path.exists(path):
        raise FileNotFoundError(f"Employee update workbook not found: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    cases = []
    for sheet_name in workbook.sheetnames:
        if not sheet_name.startswith("Update_"):
            continue
        sheet = workbook[sheet_name]
        headers = [cell.value for cell in sheet[1]]
        for values in sheet.iter_rows(min_row=2, values_only=True):
            if not any(value is not None for value in values):
                continue
            case = dict(zip(headers, values))
            case["_sheet"] = sheet_name
            cases.append(case)
    workbook.close()
    return cases


def update_employee_result(test_case_id, result, remarks=""):
    path = workbook_path()
    if not os.path.exists(path):
        return

    workbook = load_workbook(path)
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    fill = {"PASS": _PASS, "FAIL": _FAIL, "SKIPPED": _SKIP}.get(result)
    for sheet in workbook.worksheets:
        headers = [cell.value for cell in sheet[1]]
        if "Test Case ID" not in headers:
            continue
        id_column = headers.index("Test Case ID") + 1
        status_column = headers.index("Test Status") + 1 if "Test Status" in headers else None
        automation_column = headers.index("Automation Status") + 1 if "Automation Status" in headers else None
        timestamp_column = headers.index("Execution_Timestamp") + 1 if "Execution_Timestamp" in headers else None
        remarks_column = headers.index("Remarks") + 1 if "Remarks" in headers else None

        for row in range(2, sheet.max_row + 1):
            if str(sheet.cell(row, id_column).value).strip() != test_case_id:
                continue
            if status_column:
                cell = sheet.cell(row, status_column, result)
                if fill:
                    cell.fill = fill
                    cell.font = Font(bold=True)
            if automation_column:
                sheet.cell(row, automation_column, "Automated")
            if timestamp_column:
                sheet.cell(row, timestamp_column, timestamp)
            if remarks_column and remarks:
                sheet.cell(row, remarks_column, remarks[:500])
            break
    workbook.save(path)
    workbook.close()
