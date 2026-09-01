import glob
import os
import time
from datetime import datetime
from typing import Any, Dict, List, Optional

import openpyxl

from vendor_utils.logger import get_logger

logger = get_logger("excel_reader")

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEST_DATA_DIR = os.path.join(ROOT, "test_data")
CREDENTIALS_FILE = os.path.join(TEST_DATA_DIR, "credentials.xlsx")


def get_test_cases_filepath() -> str:
    """Find Create-Vendor-Management.xlsx in test_data directory."""
    target = os.path.join(TEST_DATA_DIR, "Create-Vendor-Management.xlsx")
    if os.path.exists(target):
        return target
    candidates = glob.glob(os.path.join(TEST_DATA_DIR, "*vendor*.xlsx"))
    if candidates:
        return candidates[0]
    raise FileNotFoundError(f"No vendor test cases Excel file found in: {TEST_DATA_DIR}")


def read_credentials(role: str = "Employee") -> Dict[str, str]:
    """Read login credentials for the given role from credentials.xlsx."""
    if not os.path.exists(CREDENTIALS_FILE):
        raise FileNotFoundError(f"Credentials file not found: {CREDENTIALS_FILE}")

    wb = openpyxl.load_workbook(CREDENTIALS_FILE, data_only=True)
    ws = wb["Credentials"] if "Credentials" in wb.sheetnames else wb.active
    headers = [cell.value for cell in ws[1]]

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        row_dict = dict(zip(headers, row))
        row_role = str(row_dict.get("Role", "")).strip().lower()
        if row_role == role.strip().lower() or (role.lower() in ("admin", "manager") and row_role in ("manager", "employee")):
            wb.close()
            return {
                "role": str(row_dict.get("Role", "")).strip(),
                "employee_id": str(row_dict.get("Employee_ID", "")).strip(),
                "password": str(row_dict.get("Password", "")).strip(),
                "auth_code": str(row_dict.get("Auth_Code", "111111")).strip(),
            }

    wb.close()
    # Fallback to first row
    return {"role": "Employee", "employee_id": "332", "password": "test@1234", "auth_code": "111111"}


def read_test_cases(sheet_name: str) -> List[Dict[str, Any]]:
    """Read test case rows from specified sheet in Create-Vendor-Management.xlsx."""
    file_path = get_test_cases_filepath()
    wb = openpyxl.load_workbook(file_path, data_only=True)

    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' not found in {file_path}. Available: {wb.sheetnames}")

    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]

    # Ensure required column mappings
    test_cases = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        row_dict = dict(zip(headers, row))
        tc_id = row_dict.get("Test Case ID")
        if tc_id:
            row_dict["_sheet"] = sheet_name
            # Standardize step key
            if "Test Steps" in row_dict and "Steps" not in row_dict:
                row_dict["Steps"] = row_dict["Test Steps"]
            test_cases.append(row_dict)

    wb.close()
    logger.info("Loaded %d test cases from '%s'", len(test_cases), sheet_name)
    return test_cases


def is_ui_case(tc: Dict[str, Any]) -> bool:
    """Check if test case execution type is UI."""
    exec_type = str(tc.get("Execution Type", "UI")).strip().upper()
    return "API" not in exec_type


def build_automation_id(tc_id: str) -> str:
    """Generate automation script ID from test case ID (e.g. TC_VENDOR_POS_01 -> AUT_VENDOR_POS_01)."""
    return tc_id.replace("TC_", "AUT_")


def update_test_result(
    tc_id: str,
    status: str,
    remarks: Optional[str] = None,
    sheet_name: Optional[str] = None,
) -> bool:
    """
    Update Test Status, Automation Status, Auto Script ID, and Remarks in Create-Vendor-Management.xlsx.
    Includes retry logic to safely handle open file locks.
    """
    file_path = get_test_cases_filepath()
    auto_id = build_automation_id(tc_id)
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    full_remarks = f"{timestamp} | {remarks}" if remarks else timestamp

    for attempt in range(5):
        try:
            wb = openpyxl.load_workbook(file_path)
            sheets_to_search = [sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.sheetnames

            found = False
            for sname in sheets_to_search:
                ws = wb[sname]
                headers = [cell.value for cell in ws[1]]

                # Ensure columns exist
                if "Test Case ID" not in headers:
                    continue
                tc_col = headers.index("Test Case ID") + 1

                # Locate or create columns
                def get_or_create_col(col_name: str) -> int:
                    if col_name in headers:
                        return headers.index(col_name) + 1
                    col_num = len(headers) + 1
                    ws.cell(row=1, column=col_num, value=col_name)
                    headers.append(col_name)
                    return col_num

                status_col = get_or_create_col("Test Status")
                auto_status_col = get_or_create_col("Automation Status")
                script_id_col = get_or_create_col("Auto Script ID")
                remarks_col = get_or_create_col("Remarks")

                for row_idx in range(2, ws.max_row + 1):
                    cell_val = ws.cell(row=row_idx, column=tc_col).value
                    if cell_val and str(cell_val).strip() == tc_id.strip():
                        ws.cell(row=row_idx, column=status_col, value=status)
                        ws.cell(row=row_idx, column=auto_status_col, value="Automated")
                        ws.cell(row=row_idx, column=script_id_col, value=auto_id)
                        ws.cell(row=row_idx, column=remarks_col, value=full_remarks)
                        found = True
                        break

                if found:
                    break

            if found:
                wb.save(file_path)
                wb.close()
                logger.info(
                    "Updated Excel %s: Test Status=%s, Automation Status=Automated, Auto Script ID=%s",
                    tc_id, status, auto_id,
                )
                return True
            wb.close()
            return False
        except PermissionError:
            logger.warning("Excel file locked, retrying in 0.5s (attempt %d/5)...", attempt + 1)
            time.sleep(0.5)
        except Exception as exc:
            logger.error("Failed to update Excel for %s: %s", tc_id, exc)
            return False

    return False
