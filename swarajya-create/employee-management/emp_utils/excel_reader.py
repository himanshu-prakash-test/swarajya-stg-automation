import glob
import os
from datetime import datetime
from typing import Dict, List, Optional

import openpyxl

from emp_utils.logger import get_logger

log = get_logger("excel_reader")

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(ROOT, "test_data")
CREDS_FILE = os.path.join(DATA_DIR, "credentials.xlsx")

REQUIRED_CASE_HEADERS = [
    "Test Case ID",
    "Scenario",
    "Pre Condition",
    "Steps",
    "Test Data",
    "Expected Result",
    "Auto Script ID",
    "Automation Status",
    "Execution Type",
    "Test Status",
]
RESULT_HEADERS = ["Execution Remark", "Last Run"]


def _find_cases_file() -> str:
    """Auto-discover the current test-case workbook in test_data/."""
    os.makedirs(DATA_DIR, exist_ok=True)
    patterns = [
        os.path.join(DATA_DIR, "*test-cases*.xlsx"),
        os.path.join(DATA_DIR, "*Swarajya*.xlsx"),
    ]
    for pattern in patterns:
        matches = [
            m
            for m in glob.glob(pattern)
            if "credential" not in os.path.basename(m).lower()
            and not os.path.basename(m).startswith("~$")
        ]
        if matches:
            best = max(matches, key=os.path.getmtime)
            log.info(f"Using test cases file: {os.path.basename(best)}")
            return best
    raise FileNotFoundError("No Swarajya test-case workbook found in test_data/.")


CASES_FILE = _find_cases_file()


def _clean(value) -> str:
    return str(value).strip() if value is not None else ""


def _normalise_header(header: str) -> str:
    return _clean(header).lower().replace(" ", "").replace("_", "")


def build_automation_id(tc_id: str) -> str:
    """Build the stable automation ID written back to Excel."""
    return _clean(tc_id).replace("TC_", "AUT_", 1)


def is_ui_case(tc: Dict[str, str]) -> bool:
    return _clean(tc.get("Execution Type", "UI")).upper() == "UI"


def read_credentials(role: str = "Employee") -> Dict[str, str]:
    """Pull credentials for a given role from credentials.xlsx."""
    wb = openpyxl.load_workbook(CREDS_FILE, data_only=True)
    ws = wb["Credentials"]
    headers = [_normalise_header(c.value) for c in ws[1]]
    header_map = {h: i for i, h in enumerate(headers)}

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_role = _clean(row[header_map.get("role", 0)])
        if row_role.lower() == role.lower():
            return {
                "role": row_role,
                "employee_id": _clean(row[header_map.get("employeeid", 1)]),
                "password": _clean(row[header_map.get("password", 2)]),
                "auth_code": _clean(row[header_map.get("authcode", 3)]),
            }
    raise RuntimeError(f"No credentials found for role '{role}' in {CREDS_FILE}")


def _header_map(ws) -> Dict[str, int]:
    return {
        _normalise_header(cell.value): cell.column
        for cell in ws[1]
        if _clean(cell.value)
    }


def _ensure_result_columns(ws) -> Dict[str, int]:
    header_map = _header_map(ws)
    for offset, header in enumerate(RESULT_HEADERS, start=1):
        key = _normalise_header(header)
        if key not in header_map:
            col = len(REQUIRED_CASE_HEADERS) + offset
            ws.cell(1, col, value=header)
            header_map[key] = col

    for header in REQUIRED_CASE_HEADERS:
        key = _normalise_header(header)
        if key not in header_map:
            raise RuntimeError(f"Missing required Excel header '{header}' in sheet '{ws.title}'")

    return header_map


def read_test_cases(sheet_name: str) -> List[Dict[str, str]]:
    """Load populated test-case rows from a sheet."""
    wb = openpyxl.load_workbook(CASES_FILE, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")

    ws = wb[sheet_name]
    header_map = _header_map(ws)
    missing = [h for h in REQUIRED_CASE_HEADERS if _normalise_header(h) not in header_map]
    if missing:
        raise RuntimeError(f"Sheet '{sheet_name}' is missing required headers: {missing}")

    cases = []
    tc_col = header_map["testcaseid"]
    for row_idx in range(2, ws.max_row + 1):
        tc_id = _clean(ws.cell(row_idx, tc_col).value)
        if not tc_id:
            continue
        entry = {"_sheet": sheet_name, "_row": str(row_idx)}
        for header in REQUIRED_CASE_HEADERS:
            entry[header] = _clean(ws.cell(row_idx, header_map[_normalise_header(header)]).value)
        cases.append(entry)

    log.info(f"Loaded {len(cases)} test cases from '{sheet_name}'")
    return cases


def find_test_case(tc_id: str) -> Optional[Dict[str, str]]:
    for sheet_name in ("Positive_Flows", "Negative_Flows"):
        for case in read_test_cases(sheet_name):
            if case.get("Test Case ID") == tc_id:
                return case
    return None


def update_test_result(tc_id: str, status: str, auto_id: str = "", remark: str = "") -> None:
    """Write actual result, automation status, script ID, remark, and timestamp back to Excel."""
    clean_status = _clean(status).upper()
    wb = openpyxl.load_workbook(CASES_FILE)
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for ws in wb.worksheets:
        header_map = _ensure_result_columns(ws)
        tc_col = header_map["testcaseid"]

        for row_idx in range(2, ws.max_row + 1):
            if _clean(ws.cell(row_idx, tc_col).value) != tc_id:
                continue

            execution_type = _clean(ws.cell(row_idx, header_map["executiontype"]).value).upper() or "UI"
            if execution_type == "UI":
                automation_status = "Automated"
                automation_id = auto_id or build_automation_id(tc_id)
            else:
                automation_status = "Not automated - API"
                automation_id = ""

            ws.cell(row_idx, header_map["teststatus"], value=clean_status)
            ws.cell(row_idx, header_map["automationstatus"], value=automation_status)
            ws.cell(row_idx, header_map["autoscriptid"], value=automation_id)
            ws.cell(row_idx, header_map["executionremark"], value=remark)
            for attempt in range(1, 5):
                try:
                    wb.save(CASES_FILE)
                    log.info(
                        f"Updated Excel {tc_id}: Test Status={clean_status}, "
                        f"Automation Status={automation_status}, Auto Script ID={automation_id or 'N/A'}"
                    )
                    return
                except Exception as save_err:
                    if attempt == 4:
                        raise save_err
                    import time
                    time.sleep(0.5)

    raise RuntimeError(f"Test case ID '{tc_id}' was not found in {CASES_FILE}")
