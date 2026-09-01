import os
import re
import openpyxl

class ConsultantWorkbook:
    """Parser for Create-Consultant-Management.xlsx workbook."""

    def __init__(self, file_path="consultant_mgmt/test_data/Create-Consultant-Management.xlsx"):
        self.file_path = file_path

    def _parse_field_data(self, raw_text):
        """Parse bulleted/multiline text data into a dictionary with normalized keys."""
        if not raw_text:
            return {}
        data = {}
        # Match lines like: • First Name: 'John' or • Monthly Fees: 25000 or Search Term: 'John'
        pattern = r"(?:[•\-\*]\s*)?([^:\n]+)\s*:\s*('([^']*)'|\"([^\"]*)\"|([^\n\r]+))"
        for line in str(raw_text).split("\n"):
            line = line.strip()
            if not line:
                continue
            match = re.search(pattern, line)
            if match:
                key = match.group(1).strip().lstrip("•-* ").strip()
                val = match.group(3) if match.group(3) is not None else (
                    match.group(4) if match.group(4) is not None else match.group(5)
                )
                if val is not None:
                    val = val.strip().strip("'\"")
                data[key] = val
        return data

    def _load_sheet(self, sheet_name):
        cases = []
        if not os.path.exists(self.file_path):
            return cases

        wb = openpyxl.load_workbook(self.file_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return cases

        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            tc_id = str(row[0]).strip()
            if not tc_id.startswith("TC_CONSULTANT"):
                continue

            scenario = str(row[1] or "").strip()
            precondition = str(row[2] or "").strip() if len(row) > 2 else ""
            steps = str(row[3] or "").strip() if len(row) > 3 else ""
            test_data_raw = str(row[4] or "").strip() if len(row) > 4 else ""
            expected = str(row[5] or "").strip() if len(row) > 5 else ""

            parsed_fields = self._parse_field_data(test_data_raw)

            cases.append({
                "id": tc_id,
                "type": "positive" if "POS" in tc_id else "negative",
                "scenario": scenario,
                "precondition": precondition,
                "steps": steps,
                "test_data_raw": test_data_raw,
                "fields": parsed_fields,
                "expected": expected
            })

        wb.close()
        return cases

    def get_positive_test_cases(self):
        """Extracts positive test cases from Positive_Tests sheet."""
        return self._load_sheet("Positive_Tests")

    def get_negative_test_cases(self):
        """Extracts negative test cases from Negative_Tests sheet."""
        return self._load_sheet("Negative_Tests")

    def get_all_test_cases(self):
        """Returns all 35 test cases (positive + negative)."""
        return self.get_positive_test_cases() + self.get_negative_test_cases()

    def get_test_cases(self):
        """Alias for all test cases."""
        return self.get_all_test_cases()
