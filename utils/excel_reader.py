"""
Excel utilities for the HR/Admin-only Swarajya automation.

Credentials:
    test_data/credentials.xlsx

Accepted credential headers:
    ROLE, EMPLOYEE_ID, PASSWORD, AUTH_CODE
or the repository-style:
    Role, Employee_ID, Password, Auth_Code

Test cases:
    test_data/login_test_cases.xlsx

The test-case workbook is the source of truth for the six implemented cases:
    TC_ADMIN_001..003
    TC_HR_001..003
"""

import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEST_DATA_DIR = os.path.join(BASE_DIR, "test_data")
CREDENTIALS_FILE = os.path.join(TEST_DATA_DIR, "credentials.xlsx")
TEST_CASES_FILE = os.path.join(TEST_DATA_DIR, "login_test_cases.xlsx")

PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
SKIP_FILL = PatternFill("solid", fgColor="FFEB9C")


def _normalize_header(value):
    return str(value or "").strip().lower().replace(" ", "_")


def _canonical_header(value):
    key = _normalize_header(value)
    aliases = {
        "role": "role",
        "employee_id": "employee_id",
        "employeeid": "employee_id",
        "emp_id": "employee_id",
        "password": "password",
        "auth_code": "auth_code",
        "authcode": "auth_code",
        "is_valid": "is_valid",
    }
    return aliases.get(key, key)


def read_credentials(role: str):
    if not os.path.exists(CREDENTIALS_FILE):
        raise FileNotFoundError(
            f"Missing {CREDENTIALS_FILE}. Copy your credentials workbook "
            "to test_data/credentials.xlsx before running."
        )

    wb = openpyxl.load_workbook(
        CREDENTIALS_FILE,
        read_only=True,
        data_only=True,
    )
    ws = wb.active
    headers = [c.value for c in ws[1]]
    canonical = [_canonical_header(h) for h in headers]

    for values in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(canonical, values))
        row_role = str(row.get("role") or "").strip()

        if row_role.lower() == role.lower():
            result = {
                "role": row_role,
                "employee_id": str(row.get("employee_id") or "").strip(),
                "password": str(row.get("password") or "").strip(),
                "auth_code": str(row.get("auth_code") or "").strip(),
            }
            wb.close()

            missing = [
                key for key in ("employee_id", "password", "auth_code")
                if not result[key]
            ]
            if missing:
                raise ValueError(
                    f"Missing {', '.join(missing)} for role {role} "
                    "in credentials.xlsx"
                )
            return result

    wb.close()
    raise ValueError(f"Role '{role}' was not found in credentials.xlsx")


def get_implemented_test_cases():
    """Return only the six HR/Admin cases present in the user's workbook."""
    wanted = {
        "TC_ADMIN_001", "TC_ADMIN_002", "TC_ADMIN_003",
        "TC_HR_001", "TC_HR_002", "TC_HR_003",
    }

    wb = openpyxl.load_workbook(TEST_CASES_FILE, read_only=True, data_only=True)
    cases = []

    for ws in wb.worksheets:
        headers = [c.value for c in ws[1]]
        for values in ws.iter_rows(min_row=2, values_only=True):
            row = dict(zip(headers, values))
            tc_id = str(row.get("Test Case ID") or "").strip()
            role = str(row.get("Role") or "").strip()

            if tc_id in wanted and role.lower() in {"admin", "hr"}:
                row["_sheet"] = ws.title
                cases.append(row)

    wb.close()
    return cases


def update_test_result(tc_id: str, result: str, remarks: str = ""):
    """Update the matching test case in the same Excel workbook."""
    if not os.path.exists(TEST_CASES_FILE):
        raise FileNotFoundError(TEST_CASES_FILE)

    wb = openpyxl.load_workbook(TEST_CASES_FILE)
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for ws in wb.worksheets:
        headers = [c.value for c in ws[1]]
        if "Test Case ID" not in headers:
            continue

        tc_col = headers.index("Test Case ID") + 1

        def ensure_column(name):
            nonlocal headers
            if name in headers:
                return headers.index(name) + 1
            col = ws.max_column + 1
            ws.cell(1, col, name)
            headers.append(name)
            return col

        result_col = ensure_column("Automation_Result")
        time_col = ensure_column("Execution_Timestamp")
        remarks_col = ensure_column("Remarks")

        for row_num in range(2, ws.max_row + 1):
            value = str(ws.cell(row_num, tc_col).value or "").strip()
            if value != tc_id:
                continue

            cell = ws.cell(row_num, result_col, result)
            if result == "PASS":
                cell.fill = PASS_FILL
                cell.font = Font(color="006100", bold=True)
            elif result == "FAIL":
                cell.fill = FAIL_FILL
                cell.font = Font(color="9C0006", bold=True)
            elif result == "SKIPPED":
                cell.fill = SKIP_FILL
                cell.font = Font(color="9C6500", bold=True)

            ws.cell(row_num, time_col, timestamp)
            ws.cell(row_num, remarks_col, remarks[:1000])

            if "Automation Status" in headers:
                status_col = headers.index("Automation Status") + 1
                ws.cell(row_num, status_col, "Automated")
            break

    wb.save(TEST_CASES_FILE)
    wb.close()
